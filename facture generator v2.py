import sys
import pandas as pd
from PyQt5 import QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, \
    QPushButton, QListWidget, QStackedLayout, QComboBox, QMessageBox, QMenu
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Facture Generator")
        self.setGeometry(100, 100, 800, 600)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        self.stacked_layout = QStackedLayout()
        self.layout.addLayout(self.stacked_layout)

        self.client_info = {}
        self.products = []

        self.create_client_info_page()
        self.create_product_info_page()
        self.create_payment_page()

    def create_client_info_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        self.invoice_number_input = QLineEdit()
        self.invoice_number_input.setPlaceholderText("Numéro de facture")
        layout.addWidget(self.invoice_number_input)

        self.client_name_input = QLineEdit()
        self.client_name_input.setPlaceholderText("Nom de l'entreprise client")
        layout.addWidget(self.client_name_input)

        self.client_address_input = QLineEdit()
        self.client_address_input.setPlaceholderText("Numéro et rue de l'entreprise")
        layout.addWidget(self.client_address_input)

        self.client_address_cp_input = QLineEdit()
        self.client_address_cp_input.setPlaceholderText("Code postal de l'entreprise")
        layout.addWidget(self.client_address_cp_input)

        self.client_address_ville_input = QLineEdit()
        self.client_address_ville_input.setPlaceholderText("Ville de l'entreprise")
        layout.addWidget(self.client_address_ville_input)

        self.client_contact_input = QLineEdit()
        self.client_contact_input.setPlaceholderText("Numéro de contact")
        layout.addWidget(self.client_contact_input)

        next_button = QPushButton("Page Suivante")
        next_button.clicked.connect(self.show_product_info_page)
        layout.addWidget(next_button)

        self.stacked_layout.addWidget(page)

    def create_product_info_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        self.product_name_input = QLineEdit()
        self.product_name_input.setPlaceholderText("Nom du produit")
        layout.addWidget(self.product_name_input)

        self.product_quantity_input = QLineEdit()
        self.product_quantity_input.setPlaceholderText("Quantité")
        layout.addWidget(self.product_quantity_input)

        self.product_price_input = QLineEdit()
        self.product_price_input.setPlaceholderText("Prix unitaire")
        layout.addWidget(self.product_price_input)

        add_product_button = QPushButton("Ajouter Produit")
        add_product_button.clicked.connect(self.add_product)
        layout.addWidget(add_product_button)

        self.product_list = QListWidget()
        self.product_list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.product_list.customContextMenuRequested.connect(self.show_context_menu)
        layout.addWidget(self.product_list)

        next_button = QPushButton("Page Suivante")
        next_button.clicked.connect(self.show_payment_page)
        layout.addWidget(next_button)

        self.stacked_layout.addWidget(page)

    def create_payment_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        self.payment_method_combo = QComboBox()
        self.payment_method_combo.addItem("CB")
        self.payment_method_combo.addItem("Virement")
        layout.addWidget(self.payment_method_combo)

        finish_button = QPushButton("Terminer")
        finish_button.clicked.connect(self.generate_invoice)
        layout.addWidget(finish_button)

        self.stacked_layout.addWidget(page)

    def show_product_info_page(self):
        self.client_info = {
            "Numéro de facture": self.invoice_number_input.text(),
            "Nom de l'entreprise": self.client_name_input.text(),
            "Adresse": f"{self.client_address_input.text()}, {self.client_address_cp_input.text()}, {self.client_address_ville_input.text()}",
            "Contact": self.client_contact_input.text()
        }
        self.stacked_layout.setCurrentIndex(1)

    def add_product(self):
        product_name = self.product_name_input.text()
        product_quantity = self.product_quantity_input.text()
        product_price = self.product_price_input.text()

        if not product_quantity.isdigit():
            QMessageBox.warning(self, "Erreur", "Veuillez entrer un nombre valide pour la quantité.")
            self.product_quantity_input.clear()
            return

        try:
            product_price = float(product_price)
        except ValueError:
            QMessageBox.warning(self, "Erreur", "Veuillez entrer un nombre valide pour le prix.")
            self.product_price_input.clear()
            return

        if product_name and product_quantity and product_price:
            product = {
                "Nom du produit": product_name,
                "Quantité": int(product_quantity),
                "Prix unitaire": round(product_price, 2),
                "Prix total": round(int(product_quantity) * product_price, 2)
            }
            self.products.append(product)
            self.product_list.addItem(
                f"N:{product_name} - Q:{product_quantity} - U:{product['Prix unitaire']:.2f} - T:{product['Prix total']:.2f}")
            self.product_name_input.clear()
            self.product_quantity_input.clear()
            self.product_price_input.clear()

    def show_context_menu(self, position):
        menu = QMenu()
        edit_action = menu.addAction("Modifier")
        delete_action = menu.addAction("Supprimer")

        action = menu.exec_(self.product_list.viewport().mapToGlobal(position))
        if action == edit_action:
            self.edit_product()
        elif action == delete_action:
            self.delete_product()

    def edit_product(self):
        selected_item = self.product_list.currentItem()
        if selected_item:
            product_index = self.product_list.row(selected_item)
            product = self.products[product_index]

            self.product_name_input.setText(product["Nom du produit"])
            self.product_quantity_input.setText(str(product["Quantité"]))
            self.product_price_input.setText(str(product["Prix unitaire"]))

            self.products.pop(product_index)
            self.product_list.takeItem(product_index)

    def delete_product(self):
        selected_item = self.product_list.currentItem()
        if selected_item:
            product_index = self.product_list.row(selected_item)
            self.products.pop(product_index)
            self.product_list.takeItem(product_index)

    def show_payment_page(self):
        self.stacked_layout.setCurrentIndex(2)

    def generate_invoice(self):
        payment_method = self.payment_method_combo.currentText()
        self.client_info["Mode de paiement"] = payment_method

        wb = Workbook()
        ws = wb.active
        ws.title = "Facture"

        # Add invoice number
        invoice_number = self.client_info["Numéro de facture"]
        ws.merge_cells('A1:I1') # Merge cells in one row
        cell = ws['A1']
        cell.value = f"FACTURE N° {invoice_number}"
        # Apply styles for Title
        for cell in ws["1:1"]:
            cell.font = Font(size=26, bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray background

        # Add seller info
            # Company name
        ws.merge_cells('A3:D3')
        cell = ws['A3']
        cell.value = "ELLIETECH PARIS 2014"
        cell.font = Font(bold=True)
            # Adress
        ws.merge_cells('A4:D4')
        cell = ws['A4']
        cell.value = "90 Rue de la Haie Coq Bâtiment 243,93300,Aubervilliers"
        cell.alignment = Alignment(horizontal="left", vertical="center",wrap_text=True)
        ws.row_dimensions[4].height = 45
            # N° SIRET
        ws.merge_cells('A5:D5')
        cell = ws['A5']
        cell.value = "N° SIRET: 98741912400019"
            # N° TVA
        ws.merge_cells('A6:D6')
        cell = ws['A6']
        cell.value = "N° TVA: FR 89 987419124"
            # Contact
        ws.merge_cells('A7:D7')
        cell = ws['A7']
        cell.value = "Contact: 07 54 12 06 47"

        # Add client info
            # Company name
        ws.merge_cells('F3:I3')
        cell = ws['F3']
        cell.value = f"Client : {self.client_info["Nom de l'entreprise"].upper()}"
        cell.font = Font(bold=True)
            # Adress
        ws.merge_cells('F4:I4')  # Merge cells in one row
        cell = ws['F4']
        cell.value = self.client_info["Adresse"]
        cell.alignment = Alignment(horizontal="left", vertical="center",wrap_text=True)
        ws.row_dimensions[4].height = 45
            # Contact
        ws.merge_cells('F5:I5')
        cell = ws['F5']
        cell.value = f"Tel : {self.client_info["Contact"]}"

        # Add style to the client info
            # Define border style thick = épais
        thick_border = Border(left=Side(style='thick'),
                              right=Side(style='thick'),
                              top=Side(style='thick'),
                              bottom=Side(style='thick'))
            # thin = mince
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Apply border to the outer edges of the merged cell
        ws['F3'].border = Border(top=thick_border.top, left=thick_border.left)
        ws['I3'].border = Border(top=thick_border.top, right=thick_border.right)
        ws['F5'].border = Border(bottom=thick_border.bottom, left=thick_border.left)
        ws['I5'].border = Border(bottom=thick_border.bottom, right=thick_border.right)

        # Add time info
            # facture payment time
        current_date = datetime.today().strftime('%d/%m/%Y')
        ws.merge_cells('A10:F10')
        cell = ws['A10']
        cell.value = f"Date de facturation: {current_date}"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray background
            # livraison time
        ws.merge_cells('A11:F11')
        cell = ws['A11']
        cell.value = f"Date de livraison: {current_date}"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray background
            # payment method
        ws.merge_cells('A12:F12')
        cell = ws['A12']
        cell.value = f"Mode de paiement: {payment_method}"
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray background

        # Add product info
        ws.append([])  # Empty row

        # Add headers with merged cells
        ws['A15'] = "Quantité"
        ws['A15'].alignment = Alignment(horizontal="center")
        ws['A15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray background

        ws.merge_cells('B15:E15')
        ws['B15'] = "Nom du produit"
        ws['B15'].alignment = Alignment(horizontal="center")
        ws['B15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray background

        ws.merge_cells('F15:G15')
        ws['F15'] = "Prix unitaire HT"
        ws['F15'].alignment = Alignment(horizontal="center")
        ws['F15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray background

        ws.merge_cells('H15:I15')
        ws['H15'] = "Prix total HT"
        ws['H15'].alignment = Alignment(horizontal="center")
        ws['H15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Add product data with merged cells
        total_price = 0
        row = 16
        for product in self.products:
            ws[f'A{row}'] = product["Quantité"]
            ws[f'A{row}'].alignment = Alignment(horizontal="center")
            ws[f'A{row}'].border = Border(left=thin_border.left, right=thin_border.right)

            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = product["Nom du produit"]
            ws[f'B{row}'].alignment = Alignment(horizontal="center")
            ws[f'B{row}'].border = Border(left=thin_border.left, right=thin_border.right)

            ws.merge_cells(f'F{row}:G{row}')
            ws[f'F{row}'] = product["Prix unitaire"]
            ws[f'F{row}'].alignment = Alignment(horizontal="center")
            ws[f'F{row}'].number_format = '#,##0.00 €'
            ws[f'F{row}'].border = Border(left=thin_border.left, right=thin_border.right)

            ws.merge_cells(f'H{row}:I{row}')
            ws[f'H{row}'] = product["Prix total"]
            ws[f'H{row}'].alignment = Alignment(horizontal="center")
            ws[f'H{row}'].number_format = '#,##0.00 €'
            ws[f'H{row}'].border = Border(left=thin_border.left, right=thin_border.right)

            total_price += product["Prix total"]

            row += 1

        # cell for total price
        ws.merge_cells(f'F{row + 1}:G{row + 1}')
        ws[f'F{row + 1}'].value = "Total HT"
        ws.merge_cells(f'H{row + 1}:I{row + 1}')
        ws[f'H{row + 1}'].value = total_price
        ws[f'H{row + 1}'].number_format = '#,##0.00 €'
        # cell for tva
        ws.merge_cells(f'F{row + 2}:G{row + 2}')
        ws[f'F{row + 2}'].value = "TVA 20%"
        ws.merge_cells(f'H{row + 2}:I{row + 2}')
        ws[f'H{row + 2}'].value = total_price * 0.2
        ws[f'H{row + 2}'].number_format = '#,##0.00 €'
        # cell for total ttc
        ws.merge_cells(f'F{row + 3}:G{row + 3}')
        ws[f'F{row + 3}'].value = "Total TTC"
        ws.merge_cells(f'H{row + 3}:I{row + 3}')
        ws[f'H{row + 3}'].value = total_price + total_price * 0.2
        ws[f'H{row + 3}'].number_format = '#,##0.00 €'

        #cell for the confirmation of the payment
        ws.merge_cells(f'A{row + 5}:I{row + 5}')
        ws[f'A{row + 5}'].value = f"Facture payée le {current_date} pour la somme de {total_price + total_price * 0.2} € par {payment_method}"

        # Set print settings to center the title
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 1
        ws.page_margins.bottom = 1
        ws.print_options.horizontalCentered = True

        wb.save("facture.xlsx")

        QMessageBox.information(self, "Succès", "La facture a été générée avec succès.")
        self.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())