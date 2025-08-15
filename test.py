import sys
import json
import os
from PyQt5 import QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, \
    QPushButton, QListWidget, QStackedLayout, QComboBox, QMessageBox, QMenu, QSpacerItem, QSizePolicy, QInputDialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# creation of a class ClientsWindow to manage the client list
class ClientsWindow(QWidget):
    def __init__(self, saved_clients, on_client_selected, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择客户")
        self.setGeometry(200, 200, 400, 300)

        self.saved_clients = saved_clients
        self.on_client_selected = on_client_selected

        layout = QVBoxLayout(self)

        # Add search bar
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("搜索客户名...")
        self.search_bar.textChanged.connect(self.filter_clients)
        layout.addWidget(self.search_bar)

        # Add client list
        self.clients_list = QListWidget()
        self.clients_list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.clients_list.customContextMenuRequested.connect(self.show_context_menu)
        for client in self.saved_clients:
            self.clients_list.addItem(client["Nom de l'entreprise"])
        layout.addWidget(self.clients_list)

        # Add select button
        select_button = QPushButton("选择")
        select_button.clicked.connect(self.select_client)
        layout.addWidget(select_button)

    def filter_clients(self, text):
        """Filter the client list based on the search bar input."""
        self.clients_list.clear()
        for client in self.saved_clients:
            if text.lower() in client["Nom de l'entreprise"].lower():
                self.clients_list.addItem(client["Nom de l'entreprise"])

    def show_context_menu(self, position):
        menu = QMenu()
        delete_action = menu.addAction("删除")

        action = menu.exec_(self.clients_list.viewport().mapToGlobal(position))
        if action == delete_action:
            self.delete_client()

    def delete_client(self):
        selected_row = self.clients_list.currentRow()
        if selected_row >= 0:
            reply = QMessageBox.question(self, "删除客户", "确定要删除这个客户吗？", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.saved_clients.pop(selected_row)
                self.clients_list.takeItem(selected_row)
                self.save_clients()

    def save_clients(self):
        with open('clients.json', 'w') as file:
            json.dump(self.saved_clients, file)

    def select_client(self):
        selected_row = self.clients_list.currentRow()
        if selected_row >= 0:
            selected_client = self.saved_clients[selected_row]
            self.on_client_selected(selected_client)
            self.close()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("发票生成器")
        self.setGeometry(100, 100, 800, 600)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        self.stacked_layout = QStackedLayout()
        self.layout.addLayout(self.stacked_layout)

        self.client_info = {}
        self.products = []
        self.saved_clients = []
        self.save_directory = ""

        self.load_saved_clients()
        self.load_invoice_number()
        self.create_client_info_page()
        self.create_product_info_page()
        self.create_payment_page()
#--------------------------------------------------load invoice number--------------------------------------------------
    # Load the invoice number from the JSON file
    def load_invoice_number(self):
        try:
            with open('invoiceNumber.json', 'r') as file:
                content = file.read().strip()
                if content:
                    data = json.loads(content)
                    self.invoice_number = data.get('invoice_number', 1)
                else:
                    self.invoice_number = 1
        except FileNotFoundError:
            self.invoice_number = 1

    def save_invoice_number(self):
        with open('invoiceNumber.json', 'w') as file:
            json.dump({'invoice_number': self.invoice_number}, file)


#--------------------------------------------------create each page--------------------------------------------------
    def create_client_info_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        # Add a dropdown to select saved clients
        clients_button = QPushButton("客户")
        clients_button.clicked.connect(self.show_clients_window)
        layout.addWidget(clients_button)

        layout.addWidget(QLabel("客人商店名"))
        self.client_name_input = QLineEdit()
        layout.addWidget(self.client_name_input)

        layout.addWidget(QLabel("客人商店 门牌号 和 街道名"))
        self.client_address_input = QLineEdit()
        layout.addWidget(self.client_address_input)

        layout.addWidget(QLabel("客人城市邮编"))
        self.client_address_cp_input = QLineEdit()
        layout.addWidget(self.client_address_cp_input)

        layout.addWidget(QLabel("客人城市"))
        self.client_address_ville_input = QLineEdit()
        layout.addWidget(self.client_address_ville_input)

        layout.addWidget(QLabel("客人联系方式"))
        self.client_contact_input = QLineEdit()
        layout.addWidget(self.client_contact_input)

        # Add Numéro TVA field
        layout.addWidget(QLabel("客人税号 (Numéro TVA)"))
        self.client_tva_input = QLineEdit()
        layout.addWidget(self.client_tva_input)

        button_layout = QHBoxLayout()
        prev_button = QPushButton("上一页")
        prev_button.clicked.connect(self.show_previous_page)
        button_layout.addWidget(prev_button)

        next_button = QPushButton("下一页")
        next_button.clicked.connect(self.show_product_info_page)
        button_layout.addWidget(next_button)

        layout.addLayout(button_layout)
        self.stacked_layout.addWidget(page)

    def create_product_info_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        self.product_name_input = QLineEdit()
        self.product_name_input.setPlaceholderText("产品名 (最多35个字母)")
        self.product_name_input.returnPressed.connect(self.focus_quantity_input)
        layout.addWidget(self.product_name_input)

        self.product_quantity_input = QLineEdit()
        self.product_quantity_input.setPlaceholderText("数量")
        self.product_quantity_input.returnPressed.connect(self.focus_price_input)
        layout.addWidget(self.product_quantity_input)

        self.product_price_input = QLineEdit()
        self.product_price_input.setPlaceholderText("单价")
        self.product_price_input.returnPressed.connect(self.add_product)
        layout.addWidget(self.product_price_input)

        add_product_button = QPushButton("添加产品")
        add_product_button.clicked.connect(self.add_product)
        layout.addWidget(add_product_button)

        self.product_list = QListWidget()
        self.product_list.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.product_list.customContextMenuRequested.connect(self.show_context_menu)
        layout.addWidget(self.product_list)

        button_layout = QHBoxLayout()
        prev_button = QPushButton("上一页")
        prev_button.clicked.connect(self.show_previous_page)
        button_layout.addWidget(prev_button)

        next_button = QPushButton("下一页")
        next_button.clicked.connect(self.show_payment_page)
        button_layout.addWidget(next_button)

        layout.addLayout(button_layout)
        self.stacked_layout.addWidget(page)

        # Set focus to the product name input field
        self.product_name_input.setFocus()

    def create_payment_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        # Add spacer at the top
        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))

        # Add title and combobox
        layout.addWidget(QLabel("支付方式", alignment=QtCore.Qt.AlignCenter))
        self.payment_method_combo = QComboBox()
        self.payment_method_combo.addItem("CB")
        self.payment_method_combo.addItem("Virement")
        self.payment_method_combo.setMinimumWidth(200)  # Set the minimum width of the combobox
        layout.addWidget(self.payment_method_combo, alignment=QtCore.Qt.AlignCenter)

        # Add spacer at the bottom
        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))

        button_layout = QHBoxLayout()
        prev_button = QPushButton("上一页")
        prev_button.clicked.connect(self.show_previous_page)
        button_layout.addWidget(prev_button)

        finish_button = QPushButton("生成发票")
        finish_button.clicked.connect(self.generate_invoice)
        button_layout.addWidget(finish_button)

        layout.addLayout(button_layout)
        self.stacked_layout.addWidget(page)

    #--------------------------------------------------show previous and next pages--------------------------------------------------

    def show_previous_page(self):
        current_index = self.stacked_layout.currentIndex()
        if current_index > 0:
            self.stacked_layout.setCurrentIndex(current_index - 1)

    def show_product_info_page(self):
        self.client_info = {
            "Numéro de facture": self.invoice_number,
            "Nom de l'entreprise": self.client_name_input.text(),
            "Adresse": f"{self.client_address_input.text()}, {self.client_address_cp_input.text()}, {self.client_address_ville_input.text()}",
            "Contact": self.client_contact_input.text(),
            "Num_TVA": self.client_tva_input.text()
        }
        self.stacked_layout.setCurrentIndex(1)

    def show_payment_page(self):
        self.stacked_layout.setCurrentIndex(2)

#--------------------------------------------------focus on input fields--------------------------------------------------
    # Focus on the next input (quantity) field when pressing Enter
    def focus_quantity_input(self):
        self.product_quantity_input.setFocus()
    # Focus on the next input (price) field when pressing Enter
    def focus_price_input(self):
        self.product_price_input.setFocus()


#--------------------------------------------------clients Functions--------------------------------------------------
    def load_saved_clients(self):
        try:
            with open('clients.json', 'r') as file:
                self.saved_clients = json.load(file)
        except FileNotFoundError:
            self.saved_clients = []

    def save_client_if_not_exists(self, client_info):
        try:
            with open('clients.json', 'r') as file:
                clients = json.load(file)
        except FileNotFoundError:
            clients = []

        client_names = [client["Nom de l'entreprise"].strip().lower() for client in clients]
        new_client_name = client_info["Nom de l'entreprise"].strip().lower()

        if new_client_name in client_names:
            existing_client = next(
                client for client in clients if client["Nom de l'entreprise"].strip().lower() == new_client_name
            )

            key_mapping = {
                "Adresse": "地址",
                "Adresse_cp": "邮政编码",
                "Adresse_ville": "城市",
                "Contact": "联系方式",
                "Num_TVA": "税号"
            }

            differences = []
            for key in ["Adresse", "Adresse_cp", "Adresse_ville", "Contact", "Num_TVA"]:
                if existing_client.get(key, "").strip() != client_info.get(key, "").strip():
                    chinese_key = key_mapping[key]
                    differences.append(f"{chinese_key}: {existing_client.get(key, '')} -> {client_info.get(key, '')}")

            if differences:
                message = "同客户一下信息不同 旧信息->新信息:\n" + "\n".join(
                    differences) + "\n\n请问是否替换 ?"
                reply = QMessageBox.question(self, "客户信息冲突", message, QMessageBox.Yes | QMessageBox.No)

                if reply == QMessageBox.Yes:
                    existing_client.update(client_info)
                    with open('clients.json', 'w') as file:
                        json.dump(clients, file, ensure_ascii=False, indent=4)
        else:
            clients.append(client_info)
            with open('clients.json', 'w') as file:
                json.dump(clients, file, ensure_ascii=False, indent=4)

    def show_clients_window(self):
        self.clients_window = ClientsWindow(self.saved_clients, self.fill_client_info)
        self.clients_window.show()

    def fill_client_info(self, client):
        self.client_name_input.setText(client["Nom de l'entreprise"])
        self.client_address_input.setText(client["Adresse"])
        self.client_address_cp_input.setText(client["Adresse_cp"])
        self.client_address_ville_input.setText(client["Adresse_ville"])
        self.client_contact_input.setText(client["Contact"])
        self.client_tva_input.setText(client.get("Num_TVA", ""))


#--------------------------------------------add product functions (add product page)--------------------------------------------------
    def add_product(self):
        product_name = self.product_name_input.text()
        product_quantity = self.product_quantity_input.text()
        product_price = self.product_price_input.text()

        if not product_quantity.isdigit():
            QMessageBox.warning(self, "错误", "请输入正确数量.")
            self.product_quantity_input.clear()
            return

        try:
            product_price = float(product_price)
        except ValueError:
            QMessageBox.warning(self, "错误", "请输入正确价格.")
            self.product_price_input.clear()
            return

        if product_name and product_quantity and product_price:
            product = {
                "Nom du produit": product_name,
                "Quantité": int(product_quantity),
                "Prix unitaire": product_price,
                "Prix total": int(product_quantity) * product_price
            }
            self.products.append(product)
            self.product_list.addItem(
                f"产品名: {product_name} // 数量: {product_quantity} // 单价: {product['Prix unitaire']:.2f} // 价格总和: {product['Prix total']:.2f}")
            self.product_name_input.clear()
            self.product_quantity_input.clear()
            self.product_price_input.clear()

            # Reset focus to the product name input field
            self.product_name_input.setFocus()

    # show context menu when right-clicking on the product list
    def show_context_menu(self, position):
        menu = QMenu()
        edit_action = menu.addAction("修改")
        delete_action = menu.addAction("删除")

        # Afficher le menu contextuel à la position correcte
        action = menu.exec_(self.product_list.viewport().mapToGlobal(position))
        if action == edit_action:
            self.edit_product()
        elif action == delete_action:
            self.delete_product()

    def edit_product(self):
        selected_item = self.product_list.currentItem()
        if selected_item:
            product_index = self.product_list.row(selected_item)
            product = self.products[product_index]

            self.product_name_input.setText(product["Nom du produit"])
            self.product_quantity_input.setText(str(product["Quantité"]))
            self.product_price_input.setText(str(product["Prix unitaire"]))

            self.products.pop(product_index)
            self.product_list.takeItem(product_index)

    def delete_product(self):
        selected_item = self.product_list.currentItem()
        if selected_item:
            product_index = self.product_list.row(selected_item)
            self.products.pop(product_index)
            self.product_list.takeItem(product_index)

#--------------------------------------------------generate invoice (creat excel file)--------------------------------------------------

    def generate_invoice(self):
        payment_method = self.payment_method_combo.currentText()
        self.client_info["Mode de paiement"] = payment_method

        wb = Workbook()
        ws = wb.active
        ws.title = "Facture"

        # Save client info if not already in the JSON file
        self.save_client_if_not_exists({
            "Nom de l'entreprise": self.client_name_input.text(),
            "Adresse": self.client_address_input.text(),
            "Adresse_cp": self.client_address_cp_input.text(),
            "Adresse_ville": self.client_address_ville_input.text(),
            "Contact": self.client_contact_input.text(),
            "Num_TVA": self.client_tva_input.text()
        })

        # Add invoice number
        invoice_number = self.client_info["Numéro de facture"]
        ws.merge_cells('A1:I1')
        cell = ws['A1']
        cell.value = f"FACTURE N° FA{invoice_number}"
        for cell in ws["1:1"]:
            cell.font = Font(name='Calibri', size=26, bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Add seller info
        ws.merge_cells('A3:D3')
        cell = ws['A3']
        cell.value = "ELLIETECH PARIS 2014"
        cell.font = Font(name='Calibri', bold=True)
        ws.merge_cells('A4:D4')
        cell = ws['A4']
        cell.value = "90 Rue de la Haie Coq Bâtiment 243,93300,Aubervilliers"
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[4].height = 45
        ws.merge_cells('A5:D5')
        cell = ws['A5']
        cell.value = "N° SIRET: 98741912400019"
        ws.merge_cells('A6:D6')
        cell = ws['A6']
        cell.value = "N° TVA: FR 89 987419124"
        ws.merge_cells('A7:D7')
        cell = ws['A7']
        cell.value = "Contact: 07 54 12 06 47"

        # Add client info
        ws.merge_cells('F3:I3')
        cell = ws['F3']
        cell.value = f"Client : {self.client_info['Nom de l\'entreprise'].upper()}"
        cell.font = Font(name='Calibri', bold=True)

        ws.merge_cells('F4:I4')
        cell = ws['F4']
        cell.value = self.client_info["Adresse"]
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[4].height = 45

        # Check if Num_TVA is filled
        if "Num_TVA" in self.client_info and self.client_info["Num_TVA"]:

            ws.merge_cells('F5:I5')
            cell = ws['F5']
            cell.value = f"N° TVA : {self.client_info['Num_TVA']}"

            ws.merge_cells('F6:I6')
            cell = ws['F6']
            cell.value = f"Tel : {self.client_info['Contact']}"
        else:
            ws.merge_cells('F5:I5')
            cell = ws['F5']
            cell.value = f"Tel : {self.client_info['Contact']}"

        # Add style to the client info
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                              bottom=Side(style='thick'))
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws['F3'].border = Border(top=thick_border.top, left=thick_border.left)
        ws['I3'].border = Border(top=thick_border.top, right=thick_border.right)

        # Check if Num_TVA is filled
        if "Num_TVA" in self.client_info and self.client_info["Num_TVA"]:
            ws['F6'].border = Border(bottom=thick_border.bottom, left=thick_border.left)
            ws['I6'].border = Border(bottom=thick_border.bottom, right=thick_border.right)
        else:
            ws['F5'].border = Border(bottom=thick_border.bottom, left=thick_border.left)
            ws['I5'].border = Border(bottom=thick_border.bottom, right=thick_border.right)

        # Add time info
        current_date = datetime.today().strftime('%d/%m/%Y')
        ws.merge_cells('A10:F10')
        cell = ws['A10']
        cell.value = f"Date de facturation: {current_date}"
        cell.font = Font(name='Calibri', bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.merge_cells('A11:F11')
        cell = ws['A11']
        cell.value = f"Date de livraison: {current_date}"
        cell.font = Font(name='Calibri', bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.merge_cells('A12:F12')
        cell = ws['A12']
        cell.value = f"Mode de paiement: {payment_method}"
        cell.font = Font(name='Calibri', bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Add product info
        ws.append([])  # Empty row
        ws['A15'] = "Quantité"
        ws['A15'].alignment = Alignment(horizontal="center")
        ws['A15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.merge_cells('B15:E15')
        ws['B15'] = "Nom du produit"
        ws['B15'].alignment = Alignment(horizontal="center")
        ws['B15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.merge_cells('F15:G15')
        ws['F15'] = "Prix unitaire HT"
        ws['F15'].alignment = Alignment(horizontal="center")
        ws['F15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.merge_cells('H15:I15')
        ws['H15'] = "Prix total HT"
        ws['H15'].alignment = Alignment(horizontal="center")
        ws['H15'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        total_price = 0
        row = 16
        for product in self.products:
            ws[f'A{row}'] = product["Quantité"]
            ws[f'A{row}'].alignment = Alignment(horizontal="center")
            ws[f'A{row}'].border = Border(left=thin_border.left, right=thin_border.right, bottom=thin_border.bottom)

            ws.merge_cells(f'B{row}:E{row}')
            ws[f'B{row}'] = product["Nom du produit"]
            ws[f'B{row}'].alignment = Alignment(horizontal="center")
            ws[f'B{row}'].border = Border(left=thin_border.left, bottom=thin_border.bottom)
            ws[f'C{row}'].border = Border( bottom=thin_border.bottom)
            ws[f'D{row}'].border = Border(bottom=thin_border.bottom)
            ws[f'E{row}'].border = Border(right=thin_border.right, bottom=thin_border.bottom)

            ws.merge_cells(f'F{row}:G{row}')
            ws[f'F{row}'] = product["Prix unitaire"]
            ws[f'F{row}'].alignment = Alignment(horizontal="center")
            ws[f'F{row}'].number_format = '#,##0.00 €'
            ws[f'F{row}'].border = Border(left=thin_border.left, bottom=thin_border.bottom)
            ws[f'G{row}'].border = Border(right=thin_border.right, bottom=thin_border.bottom)

            ws.merge_cells(f'H{row}:I{row}')
            ws[f'H{row}'] = product["Prix total"]
            ws[f'H{row}'].alignment = Alignment(horizontal="center")
            ws[f'H{row}'].number_format = '#,##0.00 €'
            ws[f'H{row}'].border = Border(left=thin_border.left, bottom=thin_border.bottom)
            ws[f'I{row}'].border = Border(right=thin_border.right, bottom=thin_border.bottom)

            total_price += product["Prix total"]
            row += 1

        ws.merge_cells(f'F{row + 1}:G{row + 1}')
        ws[f'F{row + 1}'].value = "Total HT"
        ws.merge_cells(f'H{row + 1}:I{row + 1}')
        ws[f'H{row + 1}'].value = f"{total_price:.2f}"
        ws[f'H{row + 1}'].number_format = '#,##0.00 €'
        # Check if Num_TVA is filled
        if "Num_TVA" in self.client_info and self.client_info["Num_TVA"]:
            # Set TVA 20% to 0
            ws.merge_cells(f'F{row + 2}:G{row + 2}')
            ws[f'F{row + 2}'].value = "TVA 20%"
            ws.merge_cells(f'H{row + 2}:I{row + 2}')
            ws[f'H{row + 2}'].value = "0.00"
            ws[f'H{row + 2}'].number_format = '#,##0.00 €'

            # Add the phrase about TVA communo code
            ws.merge_cells(f'A{row + 5}:I{row + 5}')
            ws[f'A{row + 5}'].value = "TVA communo code 123545"

            # Push the payment phrase to row + 6
            ws.merge_cells(f'A{row + 6}:I{row + 6}')
            ws[
                f'A{row + 6}'].value = f"Facture payée le {current_date} pour la somme de {total_price:.2f} € par {payment_method}"
        else:
            # Normal behavior for TVA and Total TTC
            ws.merge_cells(f'F{row + 2}:G{row + 2}')
            ws[f'F{row + 2}'].value = "TVA 20%"
            ws.merge_cells(f'H{row + 2}:I{row + 2}')
            ws[f'H{row + 2}'].value = f"{total_price * 0.2:.2f}"
            ws[f'H{row + 2}'].number_format = '#,##0.00 €'

            ws.merge_cells(f'F{row + 3}:G{row + 3}')
            ws[f'F{row + 3}'].value = "Total TTC"
            ws.merge_cells(f'H{row + 3}:I{row + 3}')
            ws[f'H{row + 3}'].value = f"{total_price + total_price * 0.2:.2f}"
            ws[f'H{row + 3}'].number_format = '#,##0.00 €'

            # Add the payment phrase in row + 5
            ws.merge_cells(f'A{row + 5}:I{row + 5}')
            ws[
                f'A{row + 5}'].value = f"Facture payée le {current_date} pour la somme de {total_price + total_price * 0.2:.2f} € par {payment_method}"

        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 1
        ws.page_margins.bottom = 1
        ws.print_options.horizontalCentered = True

        # Set font to Calibri for all cells except cell A1
        for row in ws.iter_rows():
            for cell in row:
                if cell.coordinate != 'A1':
                    cell.font = Font(name='Calibri')

        excel_file = f"FACTURE {invoice_number}.xlsx"
        wb.save(excel_file)

        # Delete the last generated file if it exists
        previous_invoice_number = invoice_number - 1
        previous_excel_file = f"FACTURE {previous_invoice_number}.xlsx"
        if os.path.exists(previous_excel_file):
            os.remove(previous_excel_file)

        QMessageBox.information(self, "成功", "发票已生成.")

        # Open the Excel file automatically
        os.startfile(excel_file)

        # Increment and save the new invoice number
        self.invoice_number += 1
        self.save_invoice_number()

        self.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Load the QSS file
    qss_file = "style.qss"
    if os.path.exists(qss_file):
        with open(qss_file, "r") as file:
            app.setStyleSheet(file.read())

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())